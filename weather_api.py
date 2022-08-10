import requests
import time
import openpyxl
from bs4 import BeautifulSoup

# Get Current Time and Date
date = time.strftime("%m/%d/%Y")
clock = time.strftime("%H:%M:%S")

#********************************** API RESULTS **********************************#
# Storage for Station List
chicago_station = 'KMDW'
ny_station = 'KNYC'

# Create the Two URLs for the API Call
chicago_url = f'https://api.weather.gov/stations/{chicago_station}/observations'
ny_url = f'https://api.weather.gov/stations/{ny_station}/observations'

# Execute the API Call
chicago_results = requests.get(chicago_url).json()
ny_results = requests.get(ny_url).json()

# Get the Current Temperature
chicago_temp_c = chicago_results['features'][0]['properties']['temperature']['value']
ny_temp_c = ny_results['features'][0]['properties']['temperature']['value']
print(f'chicago c {chicago_temp_c}')
print(f'nyc c {ny_temp_c}')

# Convert C to F
chicago_temp_f = (chicago_temp_c * (9/5)) + 32
ny_temp_f = (ny_temp_c * (9/5)) + 32
print(f'chicago f {chicago_temp_f}')
print(f'nyc f {ny_temp_f}')

#********************************** Website Results **********************************#
# Store the URLs
chicago_web_url = 'https://forecast.weather.gov/MapClick.php?lat=41.7881&lon=-87.7421'
ny_web_url = 'https://forecast.weather.gov/MapClick.php?lat=40.7823&lon=-73.9654'

# Use Requests to Get the Websites
chicago_web_results = requests.get(chicago_web_url).text
ny_web_results = requests.get(ny_web_url).text

# Soup the results and Parse
chicago_soup = BeautifulSoup(chicago_web_results, 'html.parser')
ny_soup = BeautifulSoup(ny_web_results, 'html.parser')

# Find the paragraph with the data we want
#chicago_div = chicago_soup.find('div', id='current_conditions-summary')
chicago_web_temp = chicago_soup.find('p', class_='myforecast-current-lrg')
ny_web_temp = ny_soup.find('p', class_='myforecast-current-lrg')


#********************************** Excel Results **********************************#

# Open an Excel Workbook to add the data
book = openpyxl.load_workbook('Temp_Data.xlsx')
sheet = book.active

# Insert a Row at the top so it is sorted properly
sheet.insert_rows(2)

# Insert the data from the Call
date_excel = sheet.cell(row=2, column=1)
date_excel.value = date
time_excel = sheet.cell(row=2, column=2)
time_excel.value = clock
# API RESULTS
chicago_f_excel = sheet.cell(row=2, column=3)
chicago_f_excel.value = chicago_temp_f
chicago_c_excel = sheet.cell(row=2, column=4)
chicago_c_excel.value = chicago_temp_c
ny_f_excel = sheet.cell(row=2, column=5)
ny_f_excel.value = ny_temp_f
ny_c_excel = sheet.cell(row=2, column=6)
ny_c_excel.value = ny_temp_c
# Web RESULTS
chicago_web_excel = sheet.cell(row=2, column=7)
chicago_web_excel.value = chicago_web_temp.text
ny_web_excel = sheet.cell(row=2, column=8)
ny_web_excel.value = ny_web_temp.text

# Save the Workbook
book.save('Temp_Data.xlsx')
